
from openpyxl import load_workbook
from glob import glob
from jinja2 import Template
from netaddr import *
import yaml

# Importing device configuration variables database into a dict and split multiple items and store
# as a list  (import MS Excel file format using openpyxl module)

wb = load_workbook(filename='ipsec-vpn-data.xlsx', use_iterators=True)
sheet_ipsec_config_data = wb['ipsec_config_data']

row_count = sheet_ipsec_config_data.max_row
column_count = sheet_ipsec_config_data.max_column

# print row_count
# print column_count

import_into_dict = {}
device_config_var = {}
hosts_file = []

for xl_row in range(2, row_count + 1):
    for xl_column in range(1, column_count + 1):

        # row 1 holds the configuration variable names and these names are the dictionary keys and rest of the
        # rows holds the configuration data variables

        import_into_dict[str(sheet_ipsec_config_data.cell(row=1, column=xl_column).value)] = \
            sheet_ipsec_config_data.cell(row=xl_row, column=xl_column).value

    for keys in import_into_dict:
        split_key_value = str(import_into_dict.get(keys)).split(',')
        # print value
        if len(split_key_value) > 1:
            device_config_var[keys] = str(import_into_dict.get(keys)).split(',')
        else:
            device_config_var[keys] = str(import_into_dict.get(keys))

    # print device_config_var
    # Dump Device config var into a YAML file
    yaml_output = yaml.dump(device_config_var, default_flow_style=False, explicit_start=True)
    # output += yaml.dump(definitions, default_flow_style=False, explicit_start=True)
    # print yaml_output
    with open('/Users/<homrdir>/Config_gen/%s.yml' % device_config_var['hostname'], 'w') as conf_file:
        yaml.dump(device_config_var, conf_file, default_flow_style=False, explicit_start=True)

    # Generate configuration using jinja2 templates
    vpn_policy_type = device_config_var.get('policyname')
    device_name = device_config_var['hostname']
    print device_name, vpn_policy_type
    hosts_file.append(device_config_var['hostname'])

    # Jinja2 Juniper SRX VPN template file.

    if vpn_policy_type == 'junipersrx':
        with open(glob('ipsec-junos-xl-data.j2')[0]) as conf_template:
            conf_data = conf_template.read()
        template = Template(conf_data)
        device_config = (template.render(device_config_var))
        with open(('/Users/<homrdir>/Config_gen/%s.conf' % device_name) ,'w') as conf_gen:
            conf_gen.write(device_config)
        print 'config generation is completed for device %s' % device_name
        print device_config

    # Cisco IOS VPN template file.
    elif vpn_policy_type == 'ciscoios':
        # convert Local IP /xx subnet ID into inverse mask IOS ACL configuration format
        traffic_select_v4_local_ip = device_config_var.get('v4_local_ip')
        # print traffic_select_v4_local_ip
        for ips in range(len(traffic_select_v4_local_ip)):
            traffic_select_v4_local_ip[ips] = '%s  %s' % (IPNetwork(traffic_select_v4_local_ip[ips]).network,
                                                          IPNetwork(traffic_select_v4_local_ip[ips]).hostmask)
        # print traffic_select_v4_local_ip
        device_config_var['v4_local_ip'] = traffic_select_v4_local_ip

        # convert Remote IP /xx subnet ID into inverse mask IOS ACL configuration format
        traffic_select_v4_remote_ip = device_config_var.get('v4_remote_ip')
        # print traffic_select_v4_remote_ip
        for ips in range(len(traffic_select_v4_remote_ip)):
            traffic_select_v4_remote_ip[ips] = '%s  %s' % (IPNetwork(traffic_select_v4_remote_ip[ips]).network,
                                                           IPNetwork(traffic_select_v4_remote_ip[ips]).hostmask)
        # print traffic_select_v4_remote_ip
        device_config_var['v4_remote_ip'] = traffic_select_v4_remote_ip

        # Jinja2 Configuration template rendering
        with open(glob('ipsec-ios-xl-data.j2')[0]) as conf_template:
            conf_data = conf_template.read()
        template = Template(conf_data)
        device_config = (template.render(device_config_var))
        with open(('/Users/<homrdir>/Config_gen/%s.conf' % device_name) ,'w') as conf_gen:
            conf_gen.write(device_config)
        print 'config generation is completed for device %s' % device_name
        print device_config

    # Cisco ASA VPN template file.
    elif vpn_policy_type == 'ciscoasa':
        # convert Local IP /xx subnet ID to ASA ACL configuration format
        traffic_select_v4_local_ip = device_config_var.get('v4_local_ip')
        # print traffic_select_v4_local_ip
        for ips in range(len(traffic_select_v4_local_ip)):
            traffic_select_v4_local_ip[ips] = '%s  %s' % (IPNetwork(traffic_select_v4_local_ip[ips]).network,
                                                          IPNetwork(traffic_select_v4_local_ip[ips]).netmask)
        # print traffic_select_v4_local_ip
        device_config_var['v4_local_ip'] = traffic_select_v4_local_ip

        # convert Remote IP /xx subnet ID to ASA ACL configuration format
        traffic_select_v4_remote_ip = device_config_var.get('v4_remote_ip')
        # print traffic_select_v4_remote_ip
        for ips in range(len(traffic_select_v4_remote_ip)):
            traffic_select_v4_remote_ip[ips] = '%s  %s' % (IPNetwork(traffic_select_v4_remote_ip[ips]).network,
                                                           IPNetwork(traffic_select_v4_remote_ip[ips]).netmask)
        # print traffic_select_v4_remote_ip
        device_config_var['v4_remote_ip'] = traffic_select_v4_remote_ip

        # Jinja2 Configuration template rendering
        with open(glob('ipsec-asa-xl-data.j2')[0]) as conf_template:
            conf_data = conf_template.read()
        template = Template(conf_data)
        device_config = (template.render(device_config_var))
        with open(('/Users/<homrdir>/Config_gen/%s.conf' % device_name),'w') as conf_gen:
            conf_gen.write(device_config)
        print 'config generation is completed for device %s' % device_name
        print device_config

# now write the 'hosts' file
with open('hosts', 'w+') as hosts_name:
    for h in hosts_file: hosts_name.write(h + '\n')
